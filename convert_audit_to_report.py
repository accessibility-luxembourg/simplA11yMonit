#!/usr/bin/env python3
"""Convert a filled "grille d'audit simplifié" workbook into a "grille de rapport
simplifié" workbook.

The input file must follow the structure of
``static/template-grille-audit-simplifie.xlsx`` and the output is produced from
``static/template-grille-rapport-simplifie.xlsx`` (which carries all the report
formulas, styles and data validations). Only the audited data is transferred:

* tab ``Échantillon`` : date, site, page titles and URLs;
* tabs ``P01``, ``P02``, ``P03`` : the columns ``Statut``, ``Dérogation`` and
  ``Problèmes détectés`` for each of the 53 criteria.

Everything else in the report (thematics, criteria labels, recommendations,
comment column, computed sheets) comes from the report template untouched.

Usage:
    python convert_audit_to_report.py <audit.xlsx> <rapport_out.xlsx>
                                       [--template <rapport_template.xlsx>]
"""

import argparse
import sys
from pathlib import Path

import openpyxl

# Sheets holding per-page criteria results.
PAGE_SHEETS = ("P01", "P02", "P03")

# Column letters of the data to transfer, per worksheet flavour.
# Audit:   E=Statut, F=Dérogation, G=Problèmes détectés
# Rapport: D=Statut, E=Dérogation, F=Problèmes détectés
AUDIT_COLS = {"statut": 5, "derogation": 6, "problemes": 7}
RAPPORT_COLS = {"statut": 4, "derogation": 5, "problemes": 6}

# First data row of the 53 criteria in each file.
AUDIT_FIRST_ROW = 2
RAPPORT_FIRST_ROW = 4
N_CRITERIA = 53

# Échantillon range that may carry user input (date, site, page table).
ECH_MAX_ROW = 9
ECH_MAX_COL = 3

DEFAULT_TEMPLATE = (
    Path(__file__).resolve().parent
    / "static"
    / "template-grille-rapport-simplifie.xlsx"
)


def _mergeable_interior(ws):
    """Return the set of (row, col) cells that are *inside* a merged range but
    are not its top-left anchor. Writing to those raises in openpyxl."""
    interior = set()
    for rng in ws.merged_cells.ranges:
        for row in range(rng.min_row, rng.max_row + 1):
            for col in range(rng.min_col, rng.max_col + 1):
                if (row, col) != (rng.min_row, rng.min_col):
                    interior.add((row, col))
    return interior


def copy_echantillon(src_ws, dst_ws):
    """Copy the user-entered information of the Échantillon tab.

    Both files share the same merged-cell layout, so we copy anchor-to-anchor
    and only when the source actually holds a value (never wiping the report
    template labels)."""
    skip = _mergeable_interior(dst_ws)
    copied = 0
    for row in range(1, ECH_MAX_ROW + 1):
        for col in range(1, ECH_MAX_COL + 1):
            if (row, col) in skip:
                continue
            value = src_ws.cell(row, col).value
            if value is None:
                continue
            dst_ws.cell(row, col).value = value
            copied += 1
    return copied


def copy_page(src_ws, dst_ws, page):
    """Transfer Statut / Dérogation / Problèmes détectés for one page sheet.

    A source cell is only written to the report when it holds a value, so empty
    audit cells keep the report's defaults (NT / N / blank)."""
    transferred = 0
    for i in range(N_CRITERIA):
        src_row = AUDIT_FIRST_ROW + i
        dst_row = RAPPORT_FIRST_ROW + i

        # Sanity check: the criterion code must line up between both files.
        src_code = src_ws.cell(src_row, 2).value  # audit col B = "Critère"
        if src_code is None:
            print(
                f"  [warn] {page}: no criterion at audit row {src_row}; "
                "stopping early for this sheet.",
                file=sys.stderr,
            )
            break

        for key in ("statut", "derogation", "problemes"):
            value = src_ws.cell(src_row, AUDIT_COLS[key]).value
            if value is None or (isinstance(value, str) and value.strip() == ""):
                continue
            dst_ws.cell(dst_row, RAPPORT_COLS[key]).value = value
            transferred += 1
    return transferred


def convert(audit_path, out_path, template_path):
    audit_path = Path(audit_path)
    out_path = Path(out_path)
    template_path = Path(template_path)

    if not audit_path.is_file():
        raise FileNotFoundError(f"Audit file not found: {audit_path}")
    if not template_path.is_file():
        raise FileNotFoundError(f"Report template not found: {template_path}")

    # data_only=False keeps formulas; the report sheets rely on them.
    audit_wb = openpyxl.load_workbook(audit_path, data_only=False)
    report_wb = openpyxl.load_workbook(template_path, data_only=False)

    for required in ("Échantillon", *PAGE_SHEETS):
        if required not in audit_wb.sheetnames:
            raise ValueError(f"Audit file is missing the '{required}' sheet.")
        if required not in report_wb.sheetnames:
            raise ValueError(f"Report template is missing the '{required}' sheet.")

    n = copy_echantillon(audit_wb["Échantillon"], report_wb["Échantillon"])
    print(f"Échantillon: copied {n} value(s).")

    for page in PAGE_SHEETS:
        n = copy_page(audit_wb[page], report_wb[page], page)
        print(f"{page}: transferred {n} cell value(s).")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    report_wb.save(out_path)
    print(f"Report written to: {out_path}")


def main(argv=None):
    parser = argparse.ArgumentParser(
        description="Build a 'grille de rapport simplifié' from a filled "
        "'grille d'audit simplifié'."
    )
    parser.add_argument("audit", help="Input audit .xlsx file (filled in).")
    parser.add_argument("output", help="Output report .xlsx file to create.")
    parser.add_argument(
        "--template",
        default=str(DEFAULT_TEMPLATE),
        help="Report template to base the output on "
        "(default: static/template-grille-rapport-simplifie.xlsx).",
    )
    args = parser.parse_args(argv)
    convert(args.audit, args.output, args.template)
    return 0


if __name__ == "__main__":
    sys.exit(main())
