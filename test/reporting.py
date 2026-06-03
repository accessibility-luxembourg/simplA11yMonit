import json
import locale
import re
from datetime import date
from pathlib import Path

import openpyxl

_dir = Path(__file__).parent
_config = json.loads((_dir.parent / 'config.json').read_text(encoding='utf-8'))
_TEMPLATE = _dir.parent / 'static' / 'template-grille-audit-simplifie.xlsx'
_OUT_DIR = _dir.parent / 'out'

_STATUS_ORDER = {'c': 0, 'na': 1, 'nc': 2}
_PAGE_SHEETS = ['P01', 'P02', 'P03']

# Column indices (1-based, matching Excel columns)
_COL_STATUS = 5   # E
_COL_ISSUES = 7   # H
_ROW_CRITERIA_START = 2


def _get_site_name(url):
    return re.sub(r'https?://', '', url).split('/')[0].split('?')[0].split('#')[0]


def _format_date(today, lang):
    if lang == 'fr':
        current_locale = locale.setlocale(locale.LC_TIME)
        try:
            locale.setlocale(locale.LC_TIME, 'fr_FR.UTF-8')
            formatted = today.strftime('%d %B %Y').lstrip('0')
        except locale.Error:
            months = [
                'janvier', 'février', 'mars', 'avril', 'mai', 'juin',
                'juillet', 'août', 'septembre', 'octobre', 'novembre', 'décembre',
            ]
            formatted = f"{today.day} {months[today.month - 1]} {today.year}"
        finally:
            locale.setlocale(locale.LC_TIME, current_locale)
        return formatted
    return today.strftime('%B %d, %Y')


def _format_issue(error):
    lines = [error.get('description', '')]
    if error.get('helpUrl'):
        lines.append(f"Documentation : {error['helpUrl']}")
    lines.append("Occurrences :")
    for node in error.get('nodes', []):
        summary = node.get('failureSummary', '')
        targets = ', '.join(str(t) for t in node.get('target', []))
        lines.append(f"- {summary} (Ex : {targets})")
    return '\n'.join(lines)


def gen_report(errors, pages, tested_pages, titles, lang):
    wb = openpyxl.load_workbook(str(_TEMPLATE))

    # Fill Échantillon sheet (site info and page list)
    ws_ech = wb['Échantillon']
    ws_ech['A3'] = f'Date : {_format_date(date.today(), lang)}'
    ws_ech['B4'] = _get_site_name(pages[0])
    for i, (page, title) in enumerate(zip(pages, titles), start=7):
        ws_ech.cell(row=i, column=2).value = title  # B = title
        ws_ech.cell(row=i, column=3).value = page   # C = URL

    # Fill P01/P02/P03 sheets (one per page)
    for page_idx, (page, sheet_name) in enumerate(zip(pages, _PAGE_SHEETS)):
        ws = wb[sheet_name]
        msgs = {}
        status = {}

        for error in errors:
            if error.get('url') != page:
                continue
            crit = error.get('rgaa')
            if crit is None:
                continue

            err_status = error.get('status')

            # Collect error messages for automated criteria (not N/A)
            if crit in _config['automatedCriteria'] and err_status != 'na':
                msg = _format_issue(error)
                if crit in msgs:
                    msgs[crit] += '\n\n' + msg
                else:
                    msgs[crit] = msg

            # Track highest-priority status (nc > na > c)
            if err_status:
                existing = status.get(crit)
                if existing is None or _STATUS_ORDER.get(existing, 0) < _STATUS_ORDER.get(err_status, 0):
                    status[crit] = err_status

        # Fully automated criteria with no issues are confirmed compliant
        if page in tested_pages:
            for crit in _config['fullyAutomatedCriteria']:
                if crit not in status and crit not in msgs:
                    status[crit] = 'c'

        # Always-compliant criteria are forced to C regardless of automated findings
        for crit in _config.get('alwaysCompliantCriteria', []):
            status[crit] = 'c'
            msgs.pop(crit, None)

        # Write statuses and issue descriptions to the worksheet
        for crit_idx, crit in enumerate(_config['allCriteria']):
            row = crit_idx + _ROW_CRITERIA_START
            if crit in status:
                ws.cell(row=row, column=_COL_STATUS).value = status[crit].upper()
            if crit in msgs:
                cell = ws.cell(row=row, column=_COL_ISSUES)
                cell.value = msgs[crit]
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

    site_name = _get_site_name(pages[0])
    out_path = _OUT_DIR / f'{site_name}.xlsx'
    wb.save(str(out_path))
    print(f'Report saved: {out_path}')
